import pathlib
import openpyxl
from misc.utils import print_latex
import pandas as pd
import os
import numpy as np
from postprocessing.metrics import *
import misc.datasets
import misc.constants as cs
from pathlib import Path
import misc.datasets


class ResultsDataset():
    def __init__(self, model, experiment, ph, dataset, legacy=False):
        """
        Object that compute all the performances of a given dataset for a given model and experiment and prediction horizon
        :param model: name of the model (e.g., "base")
        :param experiment: name of the experiment (e.g., "test")
        :param ph: prediction horizons in minutes (e.g., 30)
        :param dataset: name of the dataset (e.g., "ohio")
        :param legacy: used for old results without the params field in them #TODO remove
        """

        self.model = model
        self.experiment = experiment
        self.ph = ph
        self.dataset = dataset
        self.freq = np.max([misc.constants.freq, misc.datasets.datasets[dataset]["glucose_freq"]])
        self.legacy = legacy
        self.subjects = misc.datasets.datasets[self.dataset]["subjects"]

    def compute_results(self, details=False):
        """
        Loop through the subjects of the dataset, and compute the mean performances
        :return: mean of metrics, std of metrics
        """
        res = []
        for subject in self.subjects:
            res_subject = ResultsSubject(self.model, self.experiment, self.ph, self.dataset, subject,
                                         legacy=self.legacy).compute_mean_std_results()
            if details:
                print(self.dataset, subject, res_subject)

            res.append(res_subject[0])  # only the mean

        keys = list(res[0].keys())
        res = [list(res_.values()) for res_ in res]
        mean, std = np.nanmean(res, axis=0), np.nanstd(res, axis=0)
        return dict(zip(keys, mean)), dict(zip(keys, std))

    def compute_average_params(self):
        params = []
        for subject in self.subjects:
            res_subject = ResultsSubject(self.model, self.experiment, self.ph, self.dataset, subject,
                                         legacy=self.legacy)
            params.append(res_subject.params)

        return dict(zip(params[0].keys(), np.nanmean([list(_.values()) for _ in params], axis=0)))

    def to_latex(self, table="acc", model_name=None):
        """
        Format the results into a string for the paper in LATEX
        :param table: either "acc" or "cg_ega", corresponds to the table
        :param model_name: prefix of the string, name of the model
        :return:
        """
        mean, std = self.compute_results()
        if table == "p_ega":
            p_ega_keys = ["P_EGA_A+B", "P_EGA_A", "P_EGA_B", "P_EGA_C", "P_EGA_D", "P_EGA_E"]
            mean = [mean[k] * 100 for k in p_ega_keys]
            std = [std[k] * 100 for k in p_ega_keys]

        elif table == "r_ega":
            r_ega_keys = ["R_EGA_A+B", "R_EGA_A", "R_EGA_B", "R_EGA_lC", "R_EGA_uC", "R_EGA_lD", "R_EGA_uD", "R_EGA_lE",
                          "R_EGA_uC"]
            mean = [mean[k] * 100 for k in r_ega_keys]
            std = [std[k] * 100 for k in r_ega_keys]
        elif table == "cg_ega":
            cg_ega_keys = ["CG_EGA_AP_hypo", "CG_EGA_BE_hypo", "CG_EGA_EP_hypo", "CG_EGA_AP_eu", "CG_EGA_BE_eu",
                           "CG_EGA_EP_eu", "CG_EGA_AP_hyper", "CG_EGA_BE_hyper", "CG_EGA_EP_hyper"]
            mean = [mean[k] * 100 for k in cg_ega_keys]
            std = [std[k] * 100 for k in cg_ega_keys]
        elif table == "general":
            acc_keys = ["RMSE", "MAPE", "MASE", "CG_EGA_AP", "CG_EGA_BE", "CG_EGA_EP"]
            mean = [mean[k] if k not in ["CG_EGA_AP", "CG_EGA_BE", "CG_EGA_EP"] else mean[k] * 100 for k in acc_keys]
            std = [std[k] if k not in ["CG_EGA_AP", "CG_EGA_BE", "CG_EGA_EP"] else std[k] * 100 for k in acc_keys]
        elif table == "acc":
            acc_keys = ["RMSE", "MAPE", "MASE", "TG"]
            mean = [mean[k] for k in acc_keys]
            std = [std[k] for k in acc_keys]

        print_latex(mean, std, label=self.model)


class ResultsSubject():
    def __init__(self, model, experiment, ph, dataset, subject, params=None, results=None, legacy=False):
        """
        Object that compute all the performances of a given subject for a given model and experiment and prediction horizon
        :param model: name of the model (e.g., "base")
        :param experiment: name of the experiment (e.g., "test")
        :param ph: prediction horizons in minutes (e.g., 30)
        :param dataset: name of the dataset (e.g., "ohio")
        :param subject: name of the subject (e.g., "559")
        :param params: if params and results  are given, performances are directly compute on them, and both are saved into a file
        :param results: see params
        :param legacy: used for old results without the params field in them #TODO remove
        """
        self.model = model
        self.experiment = experiment
        self.ph = ph
        self.dataset = dataset
        self.subject = subject
        self.freq = np.max([misc.constants.freq, misc.datasets.datasets[dataset]["glucose_freq"]])

        if results is None and params is None:
            if not legacy:
                self.params, self.results = self.load_raw_results(legacy)
            else:
                self.results = self.load_raw_results(legacy)
        else:
            self.results = results
            self.params = params

    def load_raw_results(self, legacy=False, transfer=False):
        """
        Load the results from previous instance of ResultsSubject that has saved the them
        :param legacy: if legacy object shall  be used
        :return: (params dictionary), dataframe with ground truths and predictions
        """
        file = self.dataset + "_" + self.subject + ".npy"
        if not transfer:
            path = os.path.join(cs.path, "results", self.model, self.experiment, "ph-" + str(self.ph), file)
        else:
            path = os.path.join(cs.path, "results", self.model, self.experiment, "ph-" + str(self.ph), file)

        if not legacy:
            params, results = np.load(path, allow_pickle=True)
            dfs = []
            for result in results:
                df = pd.DataFrame(result, columns=["datetime", "y_true", "y_pred"])
                df = df.set_index("datetime")
                df = df.astype("float32")
                dfs.append(df)
            return params, dfs
        else:
            results = np.load(path, allow_pickle=True)
            dfs = []
            for result in results:
                df = pd.DataFrame(result, columns=["datetime", "y_true", "y_pred"])
                df = df.set_index("datetime")
                df = df.astype("float32")
                dfs.append(df)
            return dfs

    def save_raw_results(self):
        """
        Save the results and params
        :return:
        """
        dir = os.path.join(cs.path, "results", self.model, self.experiment, "ph-" + str(self.ph))
        Path(dir).mkdir(parents=True, exist_ok=True)

        saveable_results = np.array([res.reset_index().to_numpy() for res in self.results])

        np.save(os.path.join(dir, self.dataset + "_" + self.subject + ".npy"), [self.params, saveable_results])
        # np.save(os.path.join(dir, self.dataset + "_" + self.subject + ".npy"), np.array())

    def compute_raw_results(self, split_by_day=False):
        """
        Compute the raw metrics results for every split (or day, if split_by_day)
        :param split_by_day: wether the results are computed first by day and averaged, or averaged globally
        :return: dictionary of arrays of scores for the metrics
        """
        if split_by_day:
            results = []
            for res in self.results:
                for group in res.groupby(res.index.day):
                    results.append(group[1])
        else:
            results = self.results

        rmse_score = [rmse.RMSE(res_day) for res_day in results]
        mape_score = [mape.MAPE(res_day) for res_day in results]
        mase_score = [mase.MASE(res_day, self.ph, self.freq) for res_day in results]
        tg_score = [time_lag.time_gain(res_day, self.ph, self.freq, "mse") for res_day in results]
        cg_ega_score = np.array([cg_ega.CG_EGA(res_day, self.freq).simplified() for res_day in results])
        cg_ega_score2 = np.array([cg_ega.CG_EGA(res_day, self.freq).reduced() for res_day in results])
        p_ega_score = np.array([p_ega.P_EGA(res_day, self.freq).mean() for res_day in results])
        p_ega_a_plus_b_score = [p_ega.P_EGA(res_day, self.freq).a_plus_b() for res_day in results]
        r_ega_score = np.array([r_ega.R_EGA(res_day, self.freq).mean() for res_day in results])
        r_ega_a_plus_b_score = [r_ega.R_EGA(res_day, self.freq).a_plus_b() for res_day in results]

        return {
            "RMSE": rmse_score,
            "MAPE": mape_score,
            "MASE": mase_score,
            "TG": tg_score,
            "CG_EGA_AP": cg_ega_score2[:, 0],
            "CG_EGA_BE": cg_ega_score2[:, 1],
            "CG_EGA_EP": cg_ega_score2[:, 2],
            "CG_EGA_AP_hypo": cg_ega_score[:, 0],
            "CG_EGA_BE_hypo": cg_ega_score[:, 1],
            "CG_EGA_EP_hypo": cg_ega_score[:, 2],
            "CG_EGA_AP_eu": cg_ega_score[:, 3],
            "CG_EGA_BE_eu": cg_ega_score[:, 4],
            "CG_EGA_EP_eu": cg_ega_score[:, 5],
            "CG_EGA_AP_hyper": cg_ega_score[:, 6],
            "CG_EGA_BE_hyper": cg_ega_score[:, 7],
            "CG_EGA_EP_hyper": cg_ega_score[:, 8],
            "P_EGA_A+B": p_ega_a_plus_b_score,
            "P_EGA_A": p_ega_score[:, 0],
            "P_EGA_B": p_ega_score[:, 1],
            "P_EGA_C": p_ega_score[:, 2],
            "P_EGA_D": p_ega_score[:, 3],
            "P_EGA_E": p_ega_score[:, 4],
            "R_EGA_A+B": r_ega_a_plus_b_score,
            "R_EGA_A": r_ega_score[:, 0],
            "R_EGA_B": r_ega_score[:, 1],
            "R_EGA_uC": r_ega_score[:, 2],
            "R_EGA_lC": r_ega_score[:, 3],
            "R_EGA_uD": r_ega_score[:, 4],
            "R_EGA_lD": r_ega_score[:, 5],
            "R_EGA_uE": r_ega_score[:, 6],
            "R_EGA_lE": r_ega_score[:, 7],
        }

    def save_excel(self, file_name):
        file = os.path.join(cs.path, file_name)

        # results = self.get_results()
        misc_params = {"experiment": self.experiment, "step": 2, "dataset": self.dataset, "subject": self.subject,
                       "split": 0}

        p_params = {"P_" + x: self.params["p_coeff"][x] for x in ["A", "B", "C", "D", "E"]}
        r_params = {"R_" + x: self.params["r_coeff"][x] for x in ["A", "B", "C", "D", "E"]}
        params = {"c": self.params["coherence_factor"], **p_params, **r_params, "lr": self.params["lr"]}

        results = self.compute_mean_std_results()[0]

        data = {**misc_params, **params, **results}

        # if file not exist, create it with appropriate header
        if not pathlib.Path(file).is_file():
            wb = openpyxl.Workbook()
            wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
        else:
            wb = openpyxl.load_workbook(file)

        if not self.model in wb.sheetnames:
            wb.create_sheet(self.model)
            ws = wb[self.model]
            ws.append(list(data.keys()))
        else:
            ws = wb[self.model]

        ws.append(list(data.values()))
        wb.save(file)

    def compute_mean_std_results(self, split_by_day=False):
        """
        From the raw metrics scores, compute the mean and std
        :param split_by_day: wether the results are computed first by day and averaged, or averaged globally
        :return: mean of dictionary of metrics, std of dictionary of metrics
        """
        raw_results = self.compute_raw_results(split_by_day=split_by_day)

        mean = {key: val for key, val in zip(list(raw_results.keys()), np.nanmean(list(raw_results.values()), axis=1))}
        std = {key: val for key, val in zip(list(raw_results.keys()), np.nanstd(list(raw_results.values()), axis=1))}

        return mean, std

    def plot(self, day_number=0):
        """
        Plot a given day
        :param day_number: day (int) to plot
        :return: /
        """
        cg_ega.CG_EGA(self.results[0], self.freq).plot(day_number)

    def to_latex(self, table="acc", model_name=None):
        """
        Format the results into a string for the paper in LATEX
        :param table: either "acc" or "cg_ega", corresponds to the table
        :param model_name: prefix of the string, name of the model
        :return:
        """
        mean, std = self.compute_mean_std_results()
        if table == "p_ega":
            p_ega_keys = ["P_EGA_A+B", "P_EGA_A", "P_EGA_B", "P_EGA_C", "P_EGA_D", "P_EGA_E"]
            mean = [mean[k] * 100 for k in p_ega_keys]
            std = [std[k] * 100 for k in p_ega_keys]

        elif table == "r_ega":
            r_ega_keys = ["R_EGA_A+B", "R_EGA_A", "R_EGA_B", "R_EGA_lC", "R_EGA_uC", "R_EGA_lD", "R_EGA_uD", "R_EGA_lE",
                          "R_EGA_uC"]
            mean = [mean[k] * 100 for k in r_ega_keys]
            std = [std[k] * 100 for k in r_ega_keys]
        elif table == "cg_ega":
            cg_ega_keys = ["CG_EGA_AP", "CG_EGA_BE", "CG_EGA_EP", "CG_EGA_AP_hypo", "CG_EGA_BE_hypo", "CG_EGA_EP_hypo",
                           "CG_EGA_AP_eu", "CG_EGA_BE_eu",
                           "CG_EGA_EP_eu", "CG_EGA_AP_hyper", "CG_EGA_BE_hyper", "CG_EGA_EP_hyper"]
            mean = [mean[k] * 100 for k in cg_ega_keys]
            std = [std[k] * 100 for k in cg_ega_keys]
        elif table == "acc":
            acc_keys = ["RMSE", "MAPE", "TG"]
            mean = [mean[k] for k in acc_keys]
            std = [std[k] for k in acc_keys]

        print_latex(mean, std, label=self.subject)

from postprocessing.smoothing import smooth_results
def smooth_resultssubject(results_sbj, smoothing_params):
    res = results_sbj.results.copy()
    res = [smooth_results(res_split,smoothing_params) for res_split in res]
    return ResultsSubject(results_sbj.model, results_sbj.experiment + "_smooth", results_sbj.ph, results_sbj.dataset,
                          results_sbj.subject, results_sbj.params, res)

from postprocessing.smoothing import *
from postprocessing.results import ResultsSubject, ResultsDataset, smooth_resultssubject
import misc
def smooth_dataset_experiment(model, exp, dataset):
    smoothing = {"func": exponential_smoothing, "params": [0.85] if dataset == "idiab" else [0.85]}
    for sbj in misc.datasets.datasets[dataset]["subjects"]:
        smooth_resultssubject(ResultsSubject(model, exp, 30, dataset, sbj), smoothing).save_raw_results()